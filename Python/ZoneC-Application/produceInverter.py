import openpyxl

wb = openpyxl.load_workbook('produceSales-mini.xlsx')
sheet = wb.active

rows = sheet.max_row
columns = sheet.max_column
if rows < columns:
	temp = rows
	rows = columns
	columns = rows

print("Value is: ", columns, rows)
for i in range (1 , columns+1):
	for j in range (i, columns + 1):
		cell = sheet.cell(row = i, column = j)
		asymCell = sheet.cell(row = j, column = i)
		temp = cell.value
		cell.value = asymCell.value
		asymCell.value = temp

for i in range (columns+1, rows+1):
	for j in range (1, columns+1):
		cell = sheet.cell(row = i, column = j)
		asymCell = sheet.cell(row = j, column = i)
		temp = cell.value
		cell.value = asymCell.value
		asymCell.value = temp

wb.save('produceSales-Inverted.xlsx')
